
import re
import openpyxl
import pdfplumber
import pandas as pd
# from PyPDF2 import PdfFileWriter,PdfFileReader
import xlsxwriter
import camelot
import fitz
import os

class exceltool:
    def __init__(self):
        self.pdf_path = ''                          # 需要提取表格的pdf文件位置
        self.excel_path = ''                        # 获取内容的excel文件位置
        self.output_excel_path = ''                 # 要输出内容的Excel文件位置

    def set_execel_path(self, excelpath):           # 设置excel文件位置
        self.excel_path = excelpath

    def cellsearch(self, Keyword, strict= False):            # 遍历表格找到含有关键字为Keyword的单元格(i,j),并返回对应的单元格
        result = None
        wb = openpyxl.load_workbook(self.excel_path)
        sheet_ranges = wb[wb.sheetnames[0]]


        for i in range(1, sheet_ranges.max_column + 1):  # 遍历表格寻找指定内容
            for j in range(1, sheet_ranges.max_row + 1):
                if Keyword in str(sheet_ranges.cell(row=j, column=i).value):
                    if strict:
                        if str(sheet_ranges.cell(row=j, column=i).value) == Keyword:
                            result = sheet_ranges.cell(row=j, column=i)  # 返回含有关键词的单元格
                            break
                    else:
                        result = sheet_ranges.cell(row=j, column=i)  # 返回含有关键词的单元格
                        break

        return result

    def get_empty_cell(self, i, j):
        while self.get_cell(i, j).value:
            i += 1
        return i

    def get_cell(self, i, j):                    # 获得表格中坐标为（i,j）的单元格
        wb = openpyxl.load_workbook(self.excel_path)
        sheet_ranges = wb[wb.sheetnames[0]]
        result = sheet_ranges.cell(row=i, column=j)
        return result

    def get_table_from_pdf(self, pdfPath, pageNumber, tableselectRec, outputPath):     # 从PDF对应页中提取表格
        try:
            os.remove(outputPath)
        except:
            print('##')
        xls2 = xlsxwriter.Workbook(outputPath)                      # 生成一个空白表格
        sht1 = xls2.add_worksheet()
        xls2.close()                              # savepath:表格保存位置

        result_df = pd.DataFrame()
        
        pdf = pdfplumber.open(pdfPath)                            # readpath:PDF所在位置
        page = pdf.pages[pageNumber-1]
        page = page.crop(tableselectRec)
        table_setting = {
        "vertical_strategy": "lines", 
        "horizontal_strategy": "lines",
        "explicit_vertical_lines": [],
        "explicit_horizontal_lines": [],
        "snap_tolerance": 5,
        "join_tolerance": 3,
        "edge_min_length": 3,
        "min_words_vertical": 1,
        "min_words_horizontal": 1,
        "keep_blank_chars": False,
        "text_tolerance": 0,
        "text_x_tolerance": 0,
        "text_y_tolerance": 0,
        "intersection_tolerance": 3,
        "intersection_x_tolerance": 3,
        "intersection_y_tolerance": 3,
        }

        table = page.extract_table(table_setting)
        print(table)

        if table:
            table = [[None if x.__eq__('–') or x.__eq__('—') else x.replace(" ","") for x in y] for y in table]

            df_detail = pd.DataFrame(table[1:], columns=table[0])
            result_df = pd.concat([df_detail, result_df], ignore_index=True)
            result_df.dropna(axis=1, how='all', inplace=True)
            result_df.to_excel(excel_writer=outputPath, index=False, encoding='utf-8')
            self.set_execel_path(outputPath)








if __name__ == '__main__':

    pdf_path = 'Controller-Mutiple_p294.pdf'
    Table_save_path = r'result.xlsx'
    page_num = 293

    exe = exceltool()
    exe.get_table_from_pdf(pdf_path, Table_save_path, page_num)

    excel_path = r'excel\贴片电解电容模板.xlsx'
    result_path = r'result\贴片电解电容.xlsx'

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

# Pin端距Lmin，Lmax数值写入
    keyword = 'C'
    a = exe.cellsearch(keyword)
    i = a.row + 1
    j = a.col_idx
    L = exe.get_cell(i, j).value

    Lmin = float(L) - float(a.value[2:5])
    Lmax = float(L) + float(a.value[2:5])
    ws.cell(row=3, column=2).value = Lmin
    ws.cell(row=3, column=3).value = Lmax


# Pin长度Tmin，Tmax数值写入
    keyword = 'P'
    a = exe.cellsearch(keyword)
    i = a.row + 1
    j = a.col_idx
    P = exe.get_cell(i, j).value

    Tmin = (Lmin - (float(P) + float(a.value[2:5])))/2
    Tmax = (Lmax - (float(P) - float(a.value[2:5])))/2
    ws.cell(row=3, column=4).value = Tmin
    ws.cell(row=3, column=5).value = Tmax

# Pin宽度Wmin，Wmax数值写入
    keyword = 'R'
    a = exe.cellsearch(keyword)
    i = a.row + 1
    j = a.col_idx
    W = exe.get_cell(i, j).value

    Wmin = float(W[0:3])
    Wmax = float(W[5:8])
    ws.cell(row=3, column=6).value = Wmin
    ws.cell(row=3, column=7).value = Wmax

# 实体长宽A/B数值写入
    keyword = 'W'
    a = exe.cellsearch(keyword)
    i = a.row + 1
    j = a.col_idx
    A = exe.get_cell(i, j).value

    ws.cell(row=3, column=8).value = float(A)

# 高度H数值写入
    keyword = 'L'
    a = exe.cellsearch(keyword)
    i = a.row + 1
    j = a.col_idx
    H = exe.get_cell(i, j).value

    ws.cell(row=3, column=9).value = float(H)

    wb.save(result_path)



# 提取可编辑图片文本
#     pdf_path = r'C:\Users\HP\Desktop\PDF-Datasheet-导出表格模板-X1-2022-0726\pdf\ROHM_BD2614GSV-Z.pdf'
#
#     index = 112
#
#
#     box = [0, 0, 0, 0]
#     bugBox = [132, 534, 647, 138]
#     box[0] = int(bugBox[0] / 1.5)
#     box[1] = int(bugBox[1] / 1.5)
#     box[2] = box[0] + int(bugBox[2] / 1.5)
#     box[3] = box[1] + int(bugBox[3] / 1.5)
#
#     filt.set_pdf_path(pdf_path)
#     filt.get_figure_text(index, box[0], box[1], box[2], box[3])

    # print(filt.get_Tittle(index, box[0], box[1], box[2], box[3]))

    # bugBox = [117, 534, 676, 139]
    # box1 = [0,0,0,0]
    # box1[0] = int(bugBox[0] / 1.5)
    # box1[1] = int(bugBox[1] / 1.5)
    # box1[2] = box[0] + int(bugBox[2] / 1.5)
    # box1[3] = box[1] + int(bugBox[3] / 1.5)
    #
    # bugBox = [125, 709, 654, 150]
    # box2 = [0, 0, 0, 0]
    # box2[0] = int(bugBox[0] / 1.5)
    # box2[1] = int(bugBox[1] / 1.5)
    # box2[2] = box[0] + int(bugBox[2] / 1.5)
    # box2[3] = box[1] + int(bugBox[3] / 1.5)
    # bbox = box1 + box2
    # print(filt.get_Tittle(112, box1[0], box1[1], box1[2], box1[3]))
    # print(filt.get_Tittle(112, box2[0], box2[1], box2[2], box2[3]))
    #
    # t = filt.get_multi_tille(112 , bbox)
    # print(t)



    # image_page = page.crop(box , relative=True)
    # image_text = image_page.extract_text()
    # if image_text:
    #     print('True')
    #     print('_____________________________')
    #     print(image_text)
    #
    #     xls = xlsxwriter.Workbook('./image_text.xlsx')
    #     sht1 = xls.add_worksheet()
    #     i = 0
    #     j = 0
    #
    #     text = image_text.splitlines()
    #     for i in range(len(text)):
    #         celltext = text[i].split()
    #         print(celltext)
    #         for j in range(len(celltext)):
    #             sht1.write(i, j, celltext[j])
    #
    #     xls.close()
    # else:
    #     print('False')



    # with fitz.open(r'C:\Users\HP\Desktop\PDF-Datasheet-导出表格模板-X1-2022-0726\pdf\AIROHA_MT7682SN.pdf') as pdfDoc:
    #     page1 = pdfDoc[19]
    #     box = [0, 0, 0, 0]
    #     bugBox = [107, 710, 666, 346]
    #     box[0] = int(bugBox[0] / 1.5)
    #     box[1] = int(bugBox[1] / 1.5)
    #     box[2] = box[0] + int(bugBox[2] / 1.5)
    #     box[3] = box[1] + int(bugBox[3] / 1.5)
    #
    #     mat = fitz.Matrix(1.5, 1.5)
    #     clip = fitz.Rect(box)
    #     text = page1.getText(clip=clip)
    #     print(text)



    # pdf_path = r'C:\Users\HP\Desktop\PDF-Datasheet-导出表格模板-X1-2022-0726\pdf\DIODES_PI3DBS16412ZHEX.pdf'
    # page_num = 6
    # Table_save_path = r'result1.xlsx'
    # exe.get_table_from_pdf(pdf_path, Table_save_path, page_num)















